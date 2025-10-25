import os
import csv
import openpyxl

def read_csv_data(file_name):

    project_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    file_path = os.path.join(project_dir, file_name)

    with open(file_path, newline='', encoding='utf-8') as f:
        return list(csv.reader(f))[1:]


def load_json_data(filepath):
    with open(filepath, encoding='utf-8') as jsonfile:
        return json.load(jsonfile)


def read_excel_data(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in sheet.iter_rows(min_row=2):
        data.append({headers[i]: cell.value for i, cell in enumerate(row)})
    return data